# dropdowns.py

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# All dropdown mappings per sheet
DROPDOWN_CONFIGS = {
    "Modules": {
        "Type": ["Module", "Dashboard"]
    },
    "Table": {
        "Notes?": ["TRUE", "FALSE"],
        "Events?": ["TRUE", "FALSE"],
        "Timers?": ["TRUE", "FALSE"],
        "Delete?": ["TRUE", "FALSE"],
        "Clone?": ["TRUE", "FALSE"],
        "Hide Search?": ["TRUE", "FALSE"],
        "Web Form?": ["TRUE", "FALSE"],
        "Required?": ["TRUE", "FALSE"],
        "Auto Increment?": ["TRUE", "FALSE"],
        "Recalculate on each update ?": ["TRUE", "FALSE"],
        "Field Group Show Icon": ["TRUE", "FALSE"],
        "Security": ["None", "Readonly", "Full Restrict"]
    },
    "Dashboard": {
        "View Type": ["Empty", "List", "Calendar", "Chart", "Report Summary", "Kanban"],
        "Object Type": ["Table", "Report", "Form"],
        "Field Type": ["Field", "Static Text"],
        "Bold": ["TRUE", "FALSE"],
        "Italicize": ["TRUE", "FALSE"],
        "Hide Header?": ["TRUE", "FALSE"],
        "Hide Body?": ["TRUE", "FALSE"],
        "Bold?": ["TRUE", "FALSE"],
        "Italicize?": ["TRUE", "FALSE"],
        "Bold? - L": ["TRUE", "FALSE"],
        "Italicize? - L": ["TRUE", "FALSE"]
    },
    "Links": {
        "Link Type": ["Lookup", "LookupRestricted", "ManyToMany", "MasterDetail"],
    }
}


def inject_dropdowns(filepath, sheet_name):
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            print(f"❌ Sheet '{sheet_name}' not found.")
            return

        ws = wb[sheet_name]
        config = DROPDOWN_CONFIGS.get(sheet_name, {})
        headers = [cell.value for cell in ws[1]]

        for field, options in config.items():
            if field in headers:
                col_idx = headers.index(field) + 1
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(options)}"',
                    showDropDown=False
                )
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}1048576")

        wb.save(filepath)
        wb.close()
        print(f"✅ Dropdowns injected for '{sheet_name}' successfully.")
    except Exception as e:
        print(f"❌ Error in dropdown injection for '{sheet_name}': {e}")
