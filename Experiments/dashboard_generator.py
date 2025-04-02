import pandas as pd
import openai
import os
import json
import math
from dotenv import load_dotenv
from table_v1 import sanitize_field  # Import the sanitize function used in table.py
from dropdowns import inject_dropdowns
from utils import MetadataLoader, RawDataLoader, export_full_metadata
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

load_dotenv()

# Securely load API key
openai.api_key = os.getenv("OPENAI_API_KEY")

# Dropdown Options
VIEW_TYPES = ["Empty", "List", "Calendar", "Chart", "Report Summary", "Kanban"]
OBJECT_TYPES = ["Table", "Report", "Form"]
FIELD_TYPES = ["Field", "Static Text"]
BOOLEAN_OPTIONS = ["TRUE", "FALSE"]

# Positioning constants
HORIZONTAL_STEP = 60  # Base horizontal spacing between fields
VERTICAL_STEP = 55    # Base vertical spacing between fields
COLUMN_COUNT = 3      # Number of columns in a grid layout
INITIAL_X = 3         # Starting X position for the first column
INITIAL_Y = 5         # Starting Y position for the first row

# Common field patterns for layout ordering
ID_PATTERN = ["id", "code", "number", "key"]
NAME_PATTERN = ["name", "title", "label"]
DATE_PATTERN = ["date", "created", "modified", "time"]
AMOUNT_PATTERN = ["amount", "price", "cost", "value", "total"]
STATUS_PATTERN = ["status", "state", "condition", "phase"]
CONTACT_PATTERN = ["email", "phone", "contact", "address"]

MANDATORY_COLUMNS = [
    "Module", "Dashboard", "Element Id", "Object Type", 
    "Object Name", "View Type", "Bold", "Italicize", "Hide Header?", 
    "Hide Body?", "Field Type", "Field",
    "Bold?", "Italicize?", "Bold? - L", "Italicize? - L"
]

def extract_module_table_map(tables_df):
    module_map = {}
    for module, group in tables_df.groupby("Module"):
        tables = group["Table Name"].unique().tolist()
        module_map[module] = tables
    return module_map

class DashboardGeneratorAI:
    def __init__(self, modules_df, tables_df, user_data_summary):
        self.modules_df = modules_df
        self.tables_df = tables_df
        self.user_data_summary = user_data_summary

    def generate_prompt(self):
        modules = self.modules_df[self.modules_df["Type"] == "Module"]["Module"].tolist()
        module_table_map = json.dumps(extract_module_table_map(self.tables_df), indent=2)
        
        # Sanitize the user_data_summary before converting to JSON:
        sanitized_summary = {}
        for table, columns in self.user_data_summary.items():
            sanitized_summary[table] = [sanitize_field(col) for col in columns]
        
        table_column_map = json.dumps(sanitized_summary, indent=2)

        prompt = f"""
        You are a metadata architect for TableFlow ERP.

        🟢 Objective:
        Generate dashboard metadata rows using business context.

        🟢 Columns to Fill:
        {', '.join(MANDATORY_COLUMNS)}

        🟢 Business Rules:
        - Map **Module** directly from valid module names: {modules}.
        - Strictly follow this mapping to assign tables to their respective modules:
        {module_table_map}
        - Dashboard can follow the pattern "<Table> Dashboard" or based on business context.
        - Link **Object Name** to the actual table names.
        - **Element Id** and **Parent Element Id** should form a hierarchy (Dashboard is parent, fields/objects are children).
        - Object Type should be based on context (use: {OBJECT_TYPES}).
        - View Type should be business-friendly (choose from: {VIEW_TYPES}).
        - Bold, Italicize, Hide Header?, Hide Body? should default to FALSE but can be TRUE where appropriate.
        - **Field Type** should be "Field" for table columns.
        - **Field** should strictly match a column from the associated table.

        🟢 Dataset Context:
        Tables & Columns:
        {table_column_map}

        🟢 Output Example:
        [
            {{
                "Module": "Sales",
                "Dashboard": "Orders",
                "Element Id": "1",
                "Object Type": "Table",
                "Object Name": "Orders",
                "View Type": "List",
                "Bold": false,
                "Italicize": false,
                "Hide Header?": false,
                "Hide Body?": false,
                "Field Type": "Field",
                "Field": "Order ID"
            }}
        ]

        Repeat for all tables & relevant fields.
        """
        return prompt

    def call_openai(self):
        prompt = self.generate_prompt()
        response = openai.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are a metadata generator for dashboards."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3
        )
        return response.choices[0].message.content

    def parse_ai_response(self, ai_response):
        if "```json" in ai_response:
            ai_response = ai_response.split("```json")[1].split("```")[0].strip()
        df = pd.DataFrame(json.loads(ai_response))

        # Add fallback empty columns if missing
        for col in MANDATORY_COLUMNS:
            if col not in df.columns:
                df[col] = ""
            # Auto-fill boolean columns
            if col in ["Bold", "Italicize", "Hide Header?", "Hide Body?", "Bold?", "Italicize?", "Bold? - L", "Italicize? - L"]:
                df[col] = df[col].replace("", "FALSE").fillna("FALSE")

        return self.assign_ids(df)

    def assign_ids(self, df):
        element_counter = 1
        dashboard_map = {}
        for idx, row in df.iterrows():
            dashboard_key = f"{row['Module']}|{row['Dashboard']}"
            if dashboard_key not in dashboard_map:
                dashboard_map[dashboard_key] = element_counter
                element_counter += 1
            df.at[idx, "Element Id"] = str(dashboard_map[dashboard_key])
        return df

    def generate_dashboard_metadata(self):
        """Generate and return the dashboard metadata DataFrame in memory."""
        ai_output = self.call_openai()
        dashboard_df = self.parse_ai_response(ai_output)
        return dashboard_df

class PositionalAligner:
    def __init__(self, df):
        self.df = df
        self.module_position_map = {}  # Tracks module positions

    def apply_smart_layout(self):
        """
        Apply intelligent positioning to dashboard elements based on patterns
        from the working metadata file analysis
        """
        # Initialize positioning columns
        self.initialize_dataframe_columns()
        modules = self.df["Module"].unique()
        self.position_dashboard_panels_by_module(modules)
        self.position_dashboard_fields()
        self.generate_view_type_attributes()

    def initialize_dataframe_columns(self):
        if "PosX" not in self.df.columns:
            self.df["PosX"] = 0
        if "PosY" not in self.df.columns:
            self.df["PosY"] = 0
        if "Width" not in self.df.columns:
            self.df["Width"] = 6
        if "Height" not in self.df.columns:
            self.df["Height"] = 14
        if "PosX.1" not in self.df.columns:
            self.df["PosX.1"] = INITIAL_X
        if "PosY.1" not in self.df.columns:
            self.df["PosY.1"] = INITIAL_Y
        if "Width.1" not in self.df.columns:
            self.df["Width.1"] = 50
        if "Height.1" not in self.df.columns:
            self.df["Height.1"] = 50
        for col in ["Color", "Background Color", "Font", "Font Size", 
                    "Color.1", "Background Color.1", "Font.1", "Size",
                    "Color - L", "Background Color - L", "Font - L", "Size - L"]:
            if col not in self.df.columns:
                self.df[col] = 0

    def position_dashboard_panels_by_module(self, modules):
        for module in modules:
            module_df = self.df[self.df["Module"] == module]
            dashboards = module_df["Dashboard"].unique()
            current_x = 0
            current_y = 0
            max_height_in_row = 0
            for dashboard in dashboards:
                dashboard_elements = module_df[module_df["Dashboard"] == dashboard]
                field_count = len(dashboard_elements)
                width = 6
                if "Report" in dashboard_elements["Object Type"].values:
                    width += 3
                height = 14
                if field_count > 8:
                    height += math.ceil((field_count - 8) / 2) * 2
                if "Report Summary" in dashboard_elements["View Type"].values:
                    height += 4
                for idx in dashboard_elements.index:
                    self.df.at[idx, "PosX"] = current_x
                    self.df.at[idx, "PosY"] = current_y
                    self.df.at[idx, "Width"] = width
                    self.df.at[idx, "Height"] = height
                max_height_in_row = max(max_height_in_row, height)
                current_x += width
                if current_x >= 18:
                    current_x = 0
                    current_y += max_height_in_row + 5
                    max_height_in_row = 0

    def position_dashboard_fields(self):
        for dashboard in self.df["Dashboard"].unique():
            dashboard_df = self.df[self.df["Dashboard"] == dashboard]
            
            # Adaptive column calculation
            fields = dashboard_df["Field"].tolist()
            field_count = len(fields)
            column_count = min(max(1, math.ceil(math.sqrt(field_count))), 4)
            
            # More intelligent field priority
            def advanced_field_priority(field):
                if not isinstance(field, str):
                    return 100
                
                field_lower = field.lower().replace('-', ' ').replace('_', ' ')
                
                # Separate out the pattern lists
                id_patterns = ["id", "code", "number", "key"]
                name_patterns = ["name", "title", "label"]
                date_patterns = ["date", "created", "modified", "time"]
                status_patterns = ["status", "state", "condition", "phase"]
                amount_patterns = ["amount", "price", "cost", "value", "total"]
                contact_patterns = ["email", "phone", "contact", "address"]
                
                # Enhanced prioritization with more nuanced scoring
                priority_map = [
                    (id_patterns, -10),     # Highest priority (lowest number)
                    (name_patterns, -5),    # Second highest
                    (date_patterns, 0),     # Neutral
                    (status_patterns, 5),   # Lower priority
                    (amount_patterns, 10),  # Even lower
                    (contact_patterns, 15)  # Lowest standard priority
                ]
                
                for pattern_list, score in priority_map:
                    if any(p in field_lower for p in pattern_list):
                        return score
                
                return 20  # Default fallback
            
            # Sort fields by advanced priority
            field_indices = dashboard_df.index.tolist()
            sorted_indices = sorted(field_indices, key=lambda idx: advanced_field_priority(self.df.at[idx, "Field"]))
            
            # Adaptive positioning
            for position, idx in enumerate(sorted_indices):
                col = position % column_count
                row = position // column_count
                
                x_pos = INITIAL_X + (col * HORIZONTAL_STEP)
                y_pos = INITIAL_Y + (row * VERTICAL_STEP)
                
                self.df.at[idx, "PosX.1"] = x_pos
                self.df.at[idx, "PosY.1"] = y_pos
                
                # Smarter field sizing
                field_name = self.df.at[idx, "Field"]
                if isinstance(field_name, str):
                    normalized_field = field_name.lower().replace('-', ' ').replace('_', ' ')
                    
                    # Dynamic width and height calculation
                    field_width = 50
                    field_height = 74  # Default medium height
                    
                    # Special handling for specific field types
                    if any(pattern in normalized_field for pattern in ["description", "comment", "notes"]):
                        field_width = 100
                        field_height = 100
                    elif any(pattern in normalized_field for pattern in DATE_PATTERN):
                        field_height = 64
                    elif any(pattern in normalized_field for pattern in ID_PATTERN):
                        field_height = 50
                    
                    # Consider field length for width
                    if len(field_name) > 15:
                        field_width = min(field_width * 1.5, 100)
                    
                    self.df.at[idx, "Width.1"] = field_width
                    self.df.at[idx, "Height.1"] = field_height

    def get_field_priority(self, field_name):
        if any(pattern in field_name for pattern in ID_PATTERN):
            return 0
        elif any(pattern in field_name for pattern in NAME_PATTERN):
            return 1
        elif any(pattern in field_name for pattern in DATE_PATTERN):
            return 2
        elif any(pattern in field_name for pattern in AMOUNT_PATTERN):
            return 3
        elif any(pattern in field_name for pattern in STATUS_PATTERN):
            return 4
        elif any(pattern in field_name for pattern in CONTACT_PATTERN):
            return 5
        else:
            return 10

    def calculate_field_height(self, field_name):
        lower_field = field_name.lower()
        if any(name in lower_field for name in ["description", "comment", "notes", "address"]):
            return 84
        elif any(name in lower_field for name in DATE_PATTERN):
            return 64
        elif any(name in lower_field for name in ID_PATTERN):
            return 50
        else:
            return 74

    def generate_view_type_attributes(self):
        if "View Type Attributes" not in self.df.columns:
            self.df["View Type Attributes"] = ""
        
        grouped = self.df.groupby(['Dashboard', 'Object Name', 'View Type'])
        for (dashboard, object_name, view_type), group in grouped:
            field_names = group['Field'].dropna().astype(str).unique().tolist()
            def clean_field(field):
                return field.replace(" ", "").strip()
            fields_cleaned = [clean_field(f) for f in field_names if f.strip().lower() != "record id"]
            attr_string = "|".join([f"{f}::False::0::0" for f in fields_cleaned])
            self.df.loc[group.index, "View Type Attributes"] = attr_string

def add_dropdowns(excel_path, sheet_name="Dashboard"):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    header_values = [cell.value for cell in ws[1]]
    dropdown_mappings = {
        "View Type": VIEW_TYPES,
        "Object Type": OBJECT_TYPES,
        "Field Type": FIELD_TYPES,
        "Bold": BOOLEAN_OPTIONS,
        "Italicize": BOOLEAN_OPTIONS,
        "Hide Header?": BOOLEAN_OPTIONS,
        "Hide Body?": BOOLEAN_OPTIONS,
        "Bold?": BOOLEAN_OPTIONS,
        "Italicize?": BOOLEAN_OPTIONS,
        "Bold? - L": BOOLEAN_OPTIONS,
        "Italicize? - L": BOOLEAN_OPTIONS,
    }
    for col, options in dropdown_mappings.items():
        if col in header_values:
            col_idx = header_values.index(col) + 1
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            dv = DataValidation(type="list", formula1=f'"{",".join(options)}"', showDropDown=False)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}1048576")
    wb.save(excel_path)
    print("✅ Dropdowns injected into Excel!")

def clean_dashboard_fields(df):
    df["Dashboard"] = df["Dashboard"].astype(str).str.replace(" Dashboard", "", regex=False).str.strip()
    df["Module"] = df["Module"].astype(str).str.strip()
    df["Object Name"] = df["Object Name"].astype(str).str.strip()
    # Use the common sanitize_field to clean the Field column
    df["Field"] = df["Field"].astype(str).apply(sanitize_field).str.strip()
    if "View Type Attributes" in df.columns:
        for idx, row in df.iterrows():
            if pd.notna(row["View Type Attributes"]) and row["View Type Attributes"]:
                attrs = row["View Type Attributes"].split("|")
                cleaned_attrs = []
                for attr in attrs:
                    field_parts = attr.split("::")
                    field_name = sanitize_field(field_parts[0])
                    cleaned_attrs.append(f"{field_name}::{field_parts[1]}::{field_parts[2]}::{field_parts[3]}")
                df.at[idx, "View Type Attributes"] = "|".join(cleaned_attrs)
    return df


def main_pipeline(metadata_file, user_raw_file, output_file=None):
    # Load metadata (Modules + Tables)
    loader = MetadataLoader(metadata_file)
    if loader.load_excel() is not True:
        print(loader.load_excel())
        return None

    modules_df = loader.data.get('Modules', pd.DataFrame())
    tables_df = loader.data.get('Table', pd.DataFrame())

    # Load user raw data to extract business columns
    raw_loader = RawDataLoader(user_raw_file)
    if raw_loader.load_user_excel() is not True:
        print(raw_loader.load_user_excel())
        return None

    user_data_summary = raw_loader.summarize_user_data()
    print("✅ User Dataset Summary:", user_data_summary)

    # Generate Dashboard metadata using AI
    generator = DashboardGeneratorAI(modules_df, tables_df, user_data_summary)
    dashboard_df = generator.generate_dashboard_metadata()

    # Clean dashboard & module names
    dashboard_df = clean_dashboard_fields(dashboard_df)

    # Apply smart positional alignment
    aligner = PositionalAligner(dashboard_df)
    aligner.apply_smart_layout()

    # Update Dashboard sheet in metadata dictionary
    loader.data['Dashboard'] = dashboard_df

    # Export final metadata Excel if output_file is provided
    if output_file:
        export_full_metadata(loader.data, output_file)
        print("✅ Dashboard Metadata Excel ready!")
        # Inject dropdowns in all key sheets
        inject_dropdowns(output_file, "Modules")
        inject_dropdowns(output_file, "Table") 
        inject_dropdowns(output_file, "Dashboard")    
        inject_dropdowns(output_file, "Links")    
    # Return the dashboard DataFrame for further in-memory chaining if needed
    return dashboard_df

if __name__ == "__main__":
    metadata_file = "outputs/final_table_multi_field.xlsx"
    user_raw_file = "CRM_Food_Inventory_Sample.xlsx"
    output_file = "outputs/dashboard_final_output.xlsx"
    dashboard_df = main_pipeline(metadata_file, user_raw_file, output_file)
    print("✅ Done!")
