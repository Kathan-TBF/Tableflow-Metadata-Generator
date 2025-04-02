import pandas as pd
import openai
import json
import os
from dotenv import load_dotenv
from utils import MetadataLoader, RawDataLoader, export_full_metadata
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

load_dotenv()

# Securely load API key
openai.api_key = os.getenv("OPENAI_API_KEY")

PREDEFINEDS = """
- Data Type Options: Boolean, Currency, Date, DateTime, Decimal, Document, Image, Integer, List, Percentage, Text, Multi Select, Digital Signature, Rating, Radio Button, Assign to, Time
- Security Options: None, Readonly, Full Restrict
- Format Options:
  - Leave blank by default.
  - Only use **Custom** when specific formats are required.
  - If Custom is used, also provide **Mask** + **Validation Regex**.
- Unique Id will always be: "Record ID"
- TRUE/FALSE columns: Notes?, Events?, Delete?, Required?, Auto Increment?, Timers?, Clone?, Hide Search?, Web Form?, Recalculate on each update ?, Field Group Show Icon (Use native boolean true or false, NOT "TRUE"/"FALSE" strings)
"""

BOOLEAN_COLUMNS = [
    "Notes?", "Events?", "Timers?", "Delete?", "Clone?", "Hide Search?",
    "Web Form?", "Required?", "Auto Increment?", "Recalculate on each update ?",
    "Field Group Show Icon"
]

def sanitize_field(field):
    """Helper to sanitize single field name for export/import purposes"""
    invalid_chars = ['.', '+', '-', '*', '/', '(', ')', '[', ']', '"', '{', '}', '_']
    for char in invalid_chars:
        field = field.replace(char, '')
    return field

class TableGeneratorAI:
    def __init__(self, user_data_summary, modules_df):
        self.user_data_summary = user_data_summary
        self.modules_df = modules_df  # Inject Modules sheet for module mapping

    def generate_prompt(self):
        valid_modules = self.modules_df[self.modules_df["Type"] == "Module"]["Module"].tolist()
        table_summaries = ""
        # Inject columns for each table in prompt context
        for table, columns in self.user_data_summary.items():
            table_summaries += f"- **{table}** Columns: {columns}\n"
        
        prompt = f"""
        You are an enterprise metadata architect for the TableFlow ERP platform.
        
        🟢 Objective:
        - For each table (sheet) below, generate multiple metadata rows (one for each field).
        - Each table must be linked to one valid Module from this list: {valid_modules}.

        ⚠️ Important Rules:
        - DO NOT link to Dashboards.
        - Only use Modules where `Type = "Module"` from the provided list.
        - The AI should intelligently select the correct Module based on the table's business context.

        🟢 Metadata Rules:
        {PREDEFINEDS}

        🟢 Special Instructions:
        - Ensure every **Field** and **Display Field** strictly matches an existing column from the table's columns below.
        - **DO NOT create fields that do not exist in the table's columns.**
        - Leave **Format** blank for all fields unless using **Custom**. If Custom is used, also provide Mask + Validation Regex.
        - For the **Display Field**, prefer business-relevant columns such as "Lead Source", "Deal Name", or "Customer Name" **only if present** in the table.
        - If no such common field exists, fallback strictly to "Record ID".
        - For BOOLEAN columns (Notes?, Events?, Delete?, Required?, Auto Increment?, etc.), ONLY use native true/false (no strings).
        - Avoid empty cells. Default to empty strings ("") for optional text fields.

        🟢 Tables & Columns:
        {table_summaries}

        🟢 Output Example:
        [
            {{
                "Table Name": "Orders",
                "Notes?": true,
                "Events?": true,
                "Timers?": false,
                "Delete?": true,
                "Clone?": true,
                "Hide Search?": false,
                "Web Form?": true,
                "Display Field": "Order ID",
                "Unique Id": "Record ID",
                "Module": "Sales",
                "Field": "Order ID",
                "Data Type": "Integer",
                "Required?": true,
                "List Values": "",
                "Auto Increment?": true,
                "Auto Increment Start": "1",
                "Conditions": "",
                "Format": "",
                "Default Value": "",
                "Validation Condition": "",
                "ColorExpression": "",
                "Validation Regex": "",
                "Recalculate on each update ?": false,
                "Decimal Place": "0",
                "FieldGroupId": "",
                "Field Grpoup Type": "",
                "Field Group Show Icon": false,
                "Security": "None",
                "Role": "",
                "Click Rule": ""
            }}
        ]

        Repeat this for **ALL fields of ALL detected tables**.
        """
        return prompt

    def call_openai(self):
        prompt = self.generate_prompt()
        response = openai.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are a metadata generator for ERP tables."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.5
        )
        return response.choices[0].message.content

    def parse_ai_response(self, ai_response):
        if "```json" in ai_response:
            ai_response = ai_response.split("```json")[1].split("```")[0].strip()
        parsed_data = json.loads(ai_response)

        for row in parsed_data:
            row.setdefault("Display Field", "Record ID")
            row.setdefault("Unique Id", "Record ID")

            # Force Format blank unless "Custom"
            if row.get("Format", "").strip().lower() != "custom":
                row["Format"] = ""

            # Auto Increment Start
            if row.get("Auto Increment?", False):
                row["Auto Increment Start"] = str(row.get("Auto Increment Start") or "1")

            # Decimal Place default for numeric types
            if row.get("Data Type", "").lower() in ["decimal", "integer"]:
                row["Decimal Place"] = str(row.get("Decimal Place") or "0")
            else:
                row["Decimal Place"] = ""

            row["FieldGroupId"] = str(row.get("FieldGroupId") or "")

            for bool_col in BOOLEAN_COLUMNS:
                row[bool_col] = bool(row.get(bool_col, False))
        df = pd.DataFrame(parsed_data)
        return self.validate_fields(df)

    def validate_fields(self, df):
        sanitized_summary = {}
        # Step 1: sanitize columns inside user_data_summary
        for table_name, columns in self.user_data_summary.items():
            sanitized_summary[table_name] = [sanitize_field(c) for c in columns]
        # Step 2: Validate & export sanitized names directly
        for idx, row in df.iterrows():
            table_name = row["Table Name"]
            valid_columns = sanitized_summary.get(table_name, [])
            # Sanitize field on-the-fly for export/import system
            row_field_clean = sanitize_field(row["Field"])
            row_display_clean = sanitize_field(row["Display Field"])
            # Validate Field column
            if row_field_clean not in valid_columns:
                print(f"⚠️ Field '{row['Field']}' invalid for {table_name}. Defaulting to 'Record ID'")
                df.at[idx, "Field"] = "Record ID"
            else:
                df.at[idx, "Field"] = row_field_clean
            # Validate Display Field column
            if row_display_clean not in valid_columns and row_display_clean != "Record ID":
                df.at[idx, "Display Field"] = "Record ID"
            else:
                df.at[idx, "Display Field"] = row_display_clean
        return df

    def generate_table_metadata(self):
        """Generate and return the table metadata DataFrame in memory."""
        ai_output = self.call_openai()
        table_df = self.parse_ai_response(ai_output)
        return table_df

# 🟢 ADD DROPDOWN FUNCTION:
def add_boolean_dropdowns(excel_path, sheet_name="Table"):
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    header_values = [cell.value for cell in ws[1]]
    for col in BOOLEAN_COLUMNS:
        if col in header_values:
            col_idx = header_values.index(col) + 1
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            dv = DataValidation(type="list", formula1='"TRUE,FALSE"', showDropDown=False)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}1048576")  # Full column range except header
    wb.save(excel_path)
    print("✅ Dropdowns injected into Excel!")

def main_pipeline(metadata_file, user_raw_file, output_file=None):
    # Step 1: Load the metadata Excel (contains "Modules" sheet)
    loader = MetadataLoader(metadata_file)
    if loader.load_excel() is not True:
        print(loader.load_excel())
        return None

    modules_df = loader.data.get('Modules', pd.DataFrame())

    # Step 2: Load the user-uploaded raw Excel (actual business data tables)
    raw_loader = RawDataLoader(user_raw_file)
    if raw_loader.load_user_excel() is not True:
        print(raw_loader.load_user_excel())
        return None

    # Step 3: Summarize user-facing tables
    user_data_summary = raw_loader.summarize_user_data()
    print("✅ User Dataset Summary:", user_data_summary)

    # Step 4: Generate Table metadata using AI in-memory
    generator = TableGeneratorAI(user_data_summary, modules_df)
    table_df = generator.generate_table_metadata()

    # Update the "Table" sheet in the metadata dictionary
    loader.data['Table'] = table_df

    # Step 5: Optionally export Excel with all sheets intact if output_file is provided
    if output_file:
        export_full_metadata(loader.data, output_file)
        print("✅ Multi-field Table Metadata Excel ready!")
        add_boolean_dropdowns(output_file, "Table")
    
    # Return the table DataFrame for further processing if needed
    return table_df

if __name__ == "__main__":
    metadata_file = r"D:\Experiments\New Excel Approach\metadata_generator\outputs\module.xlsx"  # System metadata with Modules sheet
    user_raw_file = "CRM_Food_Inventory_Sample.xlsx"  # User raw dataset
    output_file = r"D:\Experiments\New Excel Approach\metadata_generator\outputs\final_table_multi_field.xlsx"
    table_df = main_pipeline(metadata_file, user_raw_file, output_file)
    print("✅ Done!")
